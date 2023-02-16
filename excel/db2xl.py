#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# -*- encoding: utf-8 -*-
import json
# Copyright (c) 2023 anqi.huang@outlook.com
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os

import dataset
import openpyxl
from openpyxl.styles import PatternFill


def dict2list(price_trend):
    list_of_keys = []
    list_of_values = []
    for key, val in dict(sorted(price_trend.items())).items():
        list_of_keys.append(key)
        list_of_values.append(val)

    return list_of_keys, list_of_values

def save(districts, file_name):
    xl_file = os.path.join(os.path.dirname(__file__), file_name + '.xlsx')

    if os.path.exists(xl_file):
        os.remove(xl_file)

    workbook = openpyxl.Workbook()
    active = workbook.active
    workbook.remove(active)

    directory = os.path.join(os.path.dirname(__file__), "../data")
    if not os.path.exists(directory):
        os.makedirs(directory)

    path = os.path.join(directory, file_name + '.db')
    database = dataset.connect('sqlite:///' + path)

    total = 0
    up = 0
    down = 0
    now = 0

    for district in districts:
        table = database[district]
        sheet = workbook.create_sheet(district)
        sheet.append(('房屋ID',
                      '区域',
                      '商圈',
                      '小区',
                      '标题',
                      '总价',
                      '单价',
                      '涨(降)幅度',
                      '价格走势',
                      '房屋户型',
                      '所在楼层',
                      '建筑面积',
                      '年代',
                      '朝向',
                      '装修',
                      '梯户比例',
                      '配备电梯',
                      '挂牌时间',
                      '上次交易',
                      '房屋交易年限',
                      '交易权属',
                      '房屋用途',
                      '关注人数',
                      '带看人数',
                      '更新时间(第一次)',
                      '更新时间',
                      '房屋url'))
        col_range = sheet.max_column
        # 首行冻结
        sheet.freeze_panes = 'A2'
        # 隐藏某些列
        sheet.column_dimensions['A'].hidden = 1
        sheet.column_dimensions['Y'].hidden = 1
        sheet.column_dimensions['Z'].hidden = 1

        for data in table.all():
            list_of_keys, list_of_values = dict2list(json.loads(data['price_trend'], object_hook=dict))
            trend = list_of_values[-1] - list_of_values[0]

            sheet.append((data['house_id'],
                          data['district'],
                          data['bizcircle'],
                          data['xiaoqu'],
                          data['title'],
                          data['total_price'],
                          data['unit_price'],
                          trend,
                          data['price_trend'],
                          data['layout'],
                          data['flood'],
                          data['building_area'],
                          data['building_year'],
                          data['orientation'],
                          data['decoration'],
                          data['house_elevator'],
                          data['elevator'],
                          data['listing_time'],
                          data['last_deal'],
                          data['deal_year'],
                          data['house_characteristics'],
                          data['land_usage'],
                          data['follow_number'],
                          data['look_number'],
                          data['crawl_time'],
                          data['update_time'],
                          data['house_url'])
                         )

            total += 1
            if trend > 0:
                up += 1
                # 涨价设置红色
                for col in range(1, col_range + 1):
                    cell = sheet.cell(sheet._current_row, col)
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
            elif trend < 0:
                down += 1
                # 降价设置绿色
                for col in range(1, col_range + 1):
                    cell = sheet.cell(sheet._current_row, col)
                    cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")
            else:
                now += 1

    workbook.save(xl_file)
    database.close()

    print("总房源 = {}(套), 涨价 = {}(套), 降价 = {}(套) , 持平 = {}(套)".format(total, up, down, now))


if __name__ == '__main__':
    districts = ["pudong", "minhang", "baoshan", "songjiang", "jiading", "qingpu"]
    file_name = 'sh-sf1a3a4a5p3-lianjia'
    save(districts, file_name)